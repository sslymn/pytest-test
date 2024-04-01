import json
import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from constants.globalConstants import *
class Test_Source:
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(BASE_URL)

    def teardown_method(self):
        self.driver.quit()

    
    
    # def readInvalidDataFromExcel():
    #      excelFile = openpyxl.load_workbook("data/invalidLogin.xlsx")
    #      sheet = excelFile["Sheet1"]
    #      rows = sheet.max_row #kaçıncı satıra kadar benim verim var
    #      data = []
    #      for i in range(2,rows+1):
    #          username = sheet.cell(i,1).value
    #          password = sheet.cell(i,2).value
    #          data.append((username,password))
    #      return data   
        

    def readInvalidDataFromJSON():
        with open(r'info/data.json', 'r') as file: # ilk baştaki r ön ek # testlerde genelde r kullanılır yani dosyayı okuma modunda açar.
            data = json.load(file) #JSON veriler Python veri yapısına dönüşür.
        return [(user['username'], user['password']) for user in data['invalid_login_users']] #tuple list oluşturur.

     
    def waitForElementVisible(self,locator,timeout=5):
        return WebDriverWait(self.driver,timeout).until(EC.visibility_of_element_located(locator))
 

    def login_info(self, username, password):
        userNameInput = self.waitForElementVisible((By.ID, USERNAME_ID))
        #userNameInput = WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located((By.ID, USERNAME_ID)))
        passwordInput = self.waitForElementVisible((By.ID, PASSWORD_ID))
        #passwordInput = WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located((By.ID, PASSWORD_ID)))    
        userNameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginBtn = WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located((By.ID, LOGINBUTTON_ID)))
        loginBtn.click()


    @pytest.mark.parametrize("username, password", readInvalidDataFromJSON())
    def test_invalid_login(self, username, password):
        userNameInput = self.waitForElementVisible((By.ID, USERNAME_ID))
        passwordInput = self.waitForElementVisible((By.ID, PASSWORD_ID))
        userNameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginButton = self.waitForElementVisible((By.ID,LOGINBUTTON_ID))
        loginButton.click()
        errorMessage =WebDriverWait(self.driver,5).until(EC.visibility_of_element_located((By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")))
        assert errorMessage.text == INVALIDLOGIN_TEXT

    def test_both(self):
        self.login_info("", "")
        errorMessage = self.driver.find_element(By.CSS_SELECTOR, "h3")
        testResult = errorMessage.text == TESTBOTH_TEXT
        assert testResult, "Hatalı"

    def test_fpassword(self):
        self.login_info("standard_user", "")
        errorMessage = self.driver.find_element(By.XPATH, "//div[@id='login_button_container']//form//h3")
        testResult = errorMessage.text == FPASSWORD_TEXT
        assert testResult, "Metin hatalı"

    def test_locked_out(self):
        self.login_info("locked_out_user", "secret_sauce")
        errorMessage = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        testResult = errorMessage.text == LOCKEDOUT_TEXT
        assert testResult, "Metin hatalı"

    def test_login_list(self):
        self.login_info("standard_user", "secret_sauce")
        WebDriverWait(self.driver, 5).until(EC.url_to_be(NURL))
        p_list = self.driver.find_elements(By.CLASS_NAME, PLISTCLASS_NAME)
        testResult = len(p_list) == 6
        assert testResult, "Ürün sayısı 6 değil"

    def test_login(self):
        self.login_info("problem_user", "secret_sauce")
        assert NURL in self.driver.current_url, "Giriş başarısız"
        print("Giriş başarılı")

    def test_logout(self):
        self.test_login()
        burger_button = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.ID, BURGERBUTTON_ID)))
        burger_button.click()
        logout_button = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.ID, LOGOUTBUTTON_ID)))
        logout_button.click()
        WebDriverWait(self.driver, 5).until(EC.url_to_be("https://www.saucedemo.com/"))
        assert "https://www.saucedemo.com/" in self.driver.current_url, "Çıkış başarısız"
        print("Çıkış başarılı")

    def test_checkout(self):
        self.login_info("standard_user", "secret_sauce")
        add_to_cart_button = self.driver.find_element(By.NAME, ADDCARTBUTTON_NAME)
        add_to_cart_button.click()
        remove_button = self.waitForElementVisible((By.NAME, REMOVEBUTTON_NAME))
        assert remove_button.text == "Remove"

# testClass = Test_Source()
# testClass.test_invalid_login()
# testClass.test_both()
# testClass.test_fpassword()
# testClass.test_login_list()
# testClass.test_login()
# testClass.test_logout()
# testClass.test_checkout()
